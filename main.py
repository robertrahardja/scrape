from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import openpyxl
import time


def get_links(driver, xpath):
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            return element.find_elements(By.TAG_NAME, "a")
        except StaleElementReferenceException:
            if attempt < max_attempts - 1:
                time.sleep(1)
                continue
            else:
                raise


# Set up the Selenium webdriver
driver = webdriver.Chrome()

# Send a request to the main page
main_url = "https://docs.aws.amazon.com/appsync/latest/APIReference/Welcome.html"
driver.get(main_url)

# Define XPath for the navigation menu
nav_xpath = (
    "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/nav[2]/div/div[2]/div/ul"
)

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "AWS AppSync API Reference"

# Get all links
links = get_links(driver, nav_xpath)
total_links = len(links)

print(f"Found {total_links} links to process.")

# Iterate over each link
for row, link in enumerate(links, start=1):
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            # Extract the URL and text from the link
            item_url = link.get_attribute("href")
            service_name = link.text.strip()

            if not item_url or not service_name:
                print(f"Skipping link {row}/{total_links}: Empty or invalid link")
                continue  # Skip empty or invalid links

            print(f"Processing link {row}/{total_links}: {service_name}")

            # Send a request to the item page
            driver.get(item_url)

            # Find the specific paragraph (first paragraph after the h1 title)
            paragraph_xpath = "//h1/following-sibling::p[1]"
            try:
                paragraph = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, paragraph_xpath))
                )
                paragraph_text = paragraph.text.strip()
                print(
                    f"  Found paragraph: {paragraph_text[:50]}..."
                )  # Print first 50 characters
            except TimeoutException:
                paragraph_text = "No paragraph found"
                print("  No paragraph found")

            # Write the data to the Excel sheet
            sheet.cell(row=row, column=1, value=service_name)
            sheet.cell(row=row, column=2, value=paragraph_text)

            break  # If successful, break out of the retry loop
        except StaleElementReferenceException:
            if attempt < max_attempts - 1:
                print(f"  Retry {attempt + 1}: StaleElementReferenceException")
                time.sleep(1)
                links = get_links(driver, nav_xpath)  # Refresh the links
                link = links[row - 1]  # Get the link again
                continue
            else:
                print(
                    f"  Failed after {max_attempts} attempts: StaleElementReferenceException"
                )
                sheet.cell(row=row, column=1, value="Error")
                sheet.cell(row=row, column=2, value="StaleElementReferenceException")

        # Navigate back to the main page
        driver.get(main_url)

# Save the Excel workbook
workbook.save("aws_appsync_api_reference.xlsx")
print("Excel file saved as aws_appsync_api_reference.xlsx")

# Close the webdriver
driver.quit()
print("WebDriver closed. Script completed.")
